import os
import re
import openpyxl
from openpyxl import load_workbook
from copy import copy

TURBLATTER_ZEITLOS_EXCEL_FILE = r'c:\Users\USER\Desktop\Coding Projects\Automated Quotations\excel_inputs_files\turblatter\turblatter_zeitlos.xlsx'
OUTPUT_DIR = r'c:\Users\USER\Desktop\Coding Projects\Automated Quotations\excel_output_files\turblatt'

os.makedirs(OUTPUT_DIR, exist_ok=True)


def safe_filename(name):
    return re.sub(r'[\\/*?:"<>|]', '_', str(name)).strip()


def copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.alignment = copy(src_cell.alignment)


wb_src = load_workbook(TURBLATTER_ZEITLOS_EXCEL_FILE)
code_sheets = [name for name in wb_src.sheetnames if 'code' in name]

# Collect all unique doors (id -> name) from the first code sheet
first_ws = wb_src[code_sheets[0]]
doors = {}
for row in first_ws.iter_rows(min_row=2, max_row=first_ws.max_row):
    door_id = row[1].value
    door_name = row[2].value
    if door_id is not None and door_id not in doors:
        doors[door_id] = door_name

print(f"Found {len(doors)} doors across {len(code_sheets)} sheets. Writing files...")

for door_id, door_name in doors.items():
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    ws_new.title = safe_filename(door_name)

    current_row = 1

    for sheet_name in code_sheets:
        ws_src = wb_src[sheet_name]

        # Section label
        ws_new.cell(row=current_row, column=1).value = sheet_name
        current_row += 1

        # Header
        header_row = list(ws_src.iter_rows(min_row=1, max_row=1))[0]
        for col_idx, src_cell in enumerate(header_row, 1):
            copy_cell(src_cell, ws_new.cell(row=current_row, column=col_idx))
        current_row += 1

        # Door rows
        for row in ws_src.iter_rows(min_row=2, max_row=ws_src.max_row):
            if row[1].value == door_id:
                for col_idx, src_cell in enumerate(row, 1):
                    copy_cell(src_cell, ws_new.cell(row=current_row, column=col_idx))
                current_row += 1

        # Blank row between sections
        current_row += 1

    # Copy column widths from first code sheet
    for col_letter, col_dim in wb_src[code_sheets[0]].column_dimensions.items():
        ws_new.column_dimensions[col_letter].width = col_dim.width

    filename = f"{door_id}_{safe_filename(door_name)}.xlsx"
    wb_new.save(os.path.join(OUTPUT_DIR, filename))
    print(f"  Saved: {filename}")

wb_src.close()
print(f"\nDone. {len(doors)} files written to {OUTPUT_DIR}")
