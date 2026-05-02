import openpyxl
from openpyxl.styles import PatternFill

# Load the workbook
wb = openpyxl.load_workbook('Standard Decora Door Prices Gefalzt.xlsx')
sheet = wb.active

# Get the range of the data
max_row = sheet.max_row
max_col = sheet.max_column

print("Cell colors (background fills):")
colors_used = set()

for row in range(1, min(max_row + 1, 20)):  # Limit to first 20 rows for brevity
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=row, column=col)
        fill = cell.fill
        if fill and fill.fill_type and fill.fill_type != 'none':
            fg = fill.fgColor
            if fg.type == 'theme':
                key = f"theme={fg.theme}, tint={round(fg.tint, 3)}"
            elif fg.type == 'rgb':
                key = f"rgb=#{fg.rgb}"
            else:
                key = None
            if key:
                colors_used.add(key)
                print(f"Row {row}, Col {col} ({cell.value}): {key}")

print(f"\nUnique colors used: {list(colors_used)}")